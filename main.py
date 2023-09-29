import os
from flask import Flask, render_template, request
import tabula
from tabula.io import read_pdf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from flask import send_file
from flask import redirect, url_for

import re

app = Flask(__name__, static_folder='static')

UPLOAD_FOLDER = os.path.join(app.root_path, 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    # Get the uploaded file from the request
    file = request.files['file']

    # Set input and output paths
    input_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
   
    output_path = os.path.join(app.root_path, 'output', "excel.xlsx")

    # Save the uploaded file
    file.save(input_path)

    # Extract tables from PDF file using tabula-py
    tables = tabula.io.read_pdf(input_path, pages='all', multiple_tables=True, encoding='latin1')

    # Create a new Excel workbook and set the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define styles for table formatting
    header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    header_border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))
    cell_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

    # Write the table headings to the Excel sheet
    for table in tables:
        heading_row = table.columns.tolist()
        ws.append(heading_row)

        # Apply formatting to the table heading
        bold_font = Font(bold=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = align_center
            cell.fill = header_fill
            cell.border = header_border

        # Write the table contents to the Excel sheet
        for row in table.values:
            ws.append(list(row))

            # Apply formatting to the table cells
            max_row = ws.max_row
            max_col = ws.max_column
            for i in range(1, max_col+1):
                column_letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[column_letter].auto_size = True
                for j in range(1, max_row+1):
                    ws.cell(row=j, column=i).alignment = align_center
                    ws.cell(row=j, column=i).border = cell_border

    # Set table margins
    table_border = Border(top=Side(style='medium', color='FF0000'), bottom=Side(style='medium', color='FF0000'), left=Side(style='medium', color='FF0000'), right=Side(style='medium', color='FF0000'))
    ws.border = table_border

    # Save the Excel workbook
    wb.save(output_path)

    return render_template('success.html', filename='excel.xlsx')

@app.route('/download/<string:filename>')
def download(filename):
    # Set the path of the generated Excel file
    output_path = os.path.join(app.root_path, 'output', filename)
    
    # Send the file as a response for download
    return send_file(output_path, as_attachment=True)
@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'GET':
        return render_template('upload.html')
    elif request.method == 'POST':
        uploaded_files = request.files.getlist('file')
        if len(uploaded_files) < 1:
            return "Please upload at least one Excel file."

        file_paths = []
        for uploaded_file in uploaded_files:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_file.filename)
            uploaded_file.save(file_path)
            file_paths.append(file_path)

        # Standard column names to be considered
        standard_columns = ['PARTICULAR', 'RATE']

        element_rates = {}
        vendor_names = {file_path: f"Vendor {i+1}" for i, file_path in enumerate(file_paths)}

        for file_path in file_paths:
            file_data = pd.read_excel(file_path, usecols=lambda x: any(col in x.upper() for col in standard_columns))
            file_data.rename(columns={col: col.capitalize() for col in standard_columns}, inplace=True)
            vendor_name = vendor_names[file_path]

            for i, row in file_data.iterrows():
                element = str(row['Particular']).strip().lower()
                rate = row['Rate']

                numeric_rate = re.search(r'\d+\.\d+', str(rate))
                if numeric_rate:
                    rate = float(numeric_rate.group())
                else:
                    rate = None

                if element not in element_rates:
                    element_rates[element] = {}
                element_rates[element][vendor_name] = rate

        final_table = pd.DataFrame.from_dict(element_rates, orient='index')
        final_table.index.name = 'Element'
        final_table['Min Rate'] = final_table.min(axis=1)
        final_table['Min Rate Vendor'] = final_table.idxmin(axis=1)
        final_table = final_table.sort_index()
        final_table = final_table.sort_values(by='Min Rate Vendor')

        output_path = os.path.join(app.root_path, 'output', 'final_table.xlsx')
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            final_table.to_excel(writer, sheet_name='Final Table', index=True)

        return render_template('success.html', filename='final_table.xlsx')


if __name__ == '__main__':
    app.run(debug=True)
